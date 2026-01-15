import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_roi_model(forecast_file, output_file):
    print(f"Loading forecast from {forecast_file}...")
    try:
        df = pd.read_csv(forecast_file)
    except FileNotFoundError:
        print("Forecast file not found. Creating dummy data.")
        df = pd.DataFrame({'category': ['Test'], 'forecast_value': [0.5]})
        
    wb = Workbook()
    ws = wb.active
    ws.title = "ROI Simulation"
    
    # 1. Setup Input Parameters Area
    ws['A1'] = "PricePulse: Inflation ROI Model"
    
    ws['A3'] = "Global Inputs"
    ws['A4'] = "Pass-through Rate (%)"
    ws['B4'] = 0.50 # Default 50%
    
    ws['A6'] = "Category Scenarios"
    
    # Headers
    headers = [
        "Category", "Current Price ($)", "Current Margin (%)", 
        "Forecasted Inflation (MoM %)", "Est. Cost Increase ($)", 
        "New Price (w/ Pass-through) ($)", "New Margin (%)", 
        "Revenue Impact (Est. %)"
    ]
    ws.append([]) # spacer
    ws.append(headers)
    
    input_start_row = 8
    
    # Aggegate forecasts per category (avg of next 3 months)
    # We really need just one row per category for the model
    stats = df.groupby('category')['forecast_value'].mean().reset_index()
    
    for idx, row in stats.iterrows():
        r = input_start_row + idx
        cat = row['category']
        inf = row['forecast_value']
        
        # Assumption Defaults
        curr_price = 100.0
        curr_margin = 0.20
        curr_cost = curr_price * (1 - curr_margin)
        
        # Excel Formulas
        # Est Cost Increase = Current Cost * (Inflation/100)
        # We use the forecasted MoM inflation as proxy for cost increase
        
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=curr_price)
        ws.cell(row=r, column=3, value=curr_margin)
        ws.cell(row=r, column=4, value=inf)
        
        # Cost Increase Formula: = Price * (1-Margin) * (Inflation/100)
        # Note: B{r} is Price, C{r} is Margin. Cost = B{r}*(1-C{r}).
        # Inflation is D{r}.
        cost_incr_formula = f"=B{r}*(1-C{r})*(D{r}/100)"
        ws.cell(row=r, column=5, value=cost_incr_formula)
        
        # New Price Formula: = Current Price + (Cost Increase * Pass_through)
        # Pass through is at B4
        new_price_formula = f"=B{r} + (E{r} * $B$4)"
        ws.cell(row=r, column=6, value=new_price_formula)
        
        # New Margin Formula: = (New Price - (Old Cost + Cost Increase)) / New Price
        # Old Cost = B{r}*(1-C{r})
        new_margin_formula = f"=(F{r} - (B{r}*(1-C{r}) + E{r})) / F{r}"
        ws.cell(row=r, column=7, value=new_margin_formula)
        
        # Rev Impact (Simple Elasticity Assumption: -0.5 elasticity)
        # % Price Change = (NewPrice - OldPrice)/OldPrice
        # % Vol Change = % Price Change * Elasticity (-1.0 for simplicity here)
        # % Rev Change = (1 + % Price Change) * (1 + % Vol Change) - 1
        # Let's just do a simple Revenue proxy
        rev_impact_formula = f"=(F{r}/B{r} - 1) * -0.5" 
        ws.cell(row=r, column=8, value=rev_impact_formula)

    # Save
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    print(f"Saved ROI model to {output_file}")

if __name__ == "__main__":
    FORECAST_FILE = "outputs/forecast.csv"
    OUTPUT_FILE = "excel_model/pricing_roi_model.xlsx"
    
    if os.path.exists(FORECAST_FILE):
        create_roi_model(FORECAST_FILE, OUTPUT_FILE)
    else:
        print(f"Error: {FORECAST_FILE} not found.")
